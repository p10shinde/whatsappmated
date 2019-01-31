from flask import send_file, Flask, send_from_directory, render_template, jsonify
from flask_api import status
from flask_cors import CORS
import time

from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException

import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import configparser
import openpyxl 
config = configparser.ConfigParser()
config.read('config.ini')
config.sections()

import base64
app = Flask(__name__)
CORS(app)

# users_ref = db.collection('users').document('userid1').get()
# print(users_ref.get('isActive'))
# db.collection('users').document('userid1').set({'isActive' : True}, merge=True)

# def addDriverListener():
# 	while True:
# 		try : 
# 			driver.find_element_by_tag_name('body')
# 			print("Its' good")
# 		except InvalidSessionIdException :
# 			print('InvalidSessionIdException')
# 		except WebDriverException : 
# 			print('WebDriverException')
# 		except Exception as e :
# 			print('#####################failed')
# 		time.sleep(2)

def firebaseSetup():
	global db 
	# Use a service account
	cred = credentials.Certificate('./wtsapproject-cd387d40d4a1.json')
	firebase_admin.initialize_app(cred)
	db = firestore.client()


# @app.route('/image')
# def get_image():
# 	filename = 'qr_code.png'
# 	with open(filename, "rb") as image_file:
# 		# data = image_file.read()
# 		# encoded_string = data.encode("base64")
# 		encoded_string = base64.b64encode(image_file.read())
# 	return encoded_string

@app.route('/start')
def startWsp():
	try :
		options = Options()
		# options.headless = True
		global driver
		driver = webdriver.Firefox(options=options, executable_path="geckodriver.exe")
		# driver.implicitly_wait(5)
		print("Firefox Headless Browser Invoked")
		driver.get('https://web.whatsapp.com')
		# addDriverListener()
		return jsonify({'msg':'Connected','data':''}), 200
	except Exception as e:
		return jsonify({'msg':'Exception','data':str(e)}), 404

@app.route('/scan')
def scanQR():
	qr_base64 = ""
	try:
		driver
	except Exception as e:
		return jsonify({'msg':'Please connect to whatsapp first.','data' : ''}), 404
	
	try:
		#before scanning always check if realod qr is visible
		# reload_qr = WebDriverWait(driver, 1).until(
		# expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "._2zblx")))
		driver.find_element_by_css_selector("._2zblx").click()
		# if reload_qr :
		driver.refresh()
		return jsonify({'msg':'try again', 'data' : 'NULL'}), 200
		
	except Exception as e:
		try:
			#qr_code image
			print('waiting ')
			qr_image_tag = WebDriverWait(driver, 10).until(
			expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "._2EZ_m img")))
			# qr_image_tag = driver.find_element_by_css_selector("._2EZ_m img")
			qr_base64 = qr_image_tag.get_attribute("src")
			# driver.get_screenshot_as_file("qr_code.png")
			print('wait done')
			# print('found :' + qr_base64)
			# encoded_string = get_image()

			 # org = driver.find_element_by_id('org')

		except Exception as e: 
			print('scan qr error')
			print(e)
			return jsonify({'msg':'Exception found ','data' :e}), 404

		
		
	# return qr_base64
	return jsonify({'msg': '', 'data' : qr_base64}), 200
		
@app.route('/isLoggedIn')
def isLoggedIn():
	try:
		# contact list locator
		driver.find_element_by_css_selector(".RLfQR")
		return jsonify({'msg':'Logged In...','data':''}), 200
	except NoSuchElementException:
		return jsonify({'msg':'Exception','data':''}), 404
	except Exception as e:
		return jsonify({'msg':'Exception','data':str(e)}), 404

@app.route('/sendMsg')
def sendMsg():
	get_data_from_excel()



def get_data_from_excel():
	wb_obj = openpyxl.load_workbook("./input/data.xlsx")
	sheet_obj = wb_obj.active
	max_row = sheet_obj.max_row
	max_col = sheet_obj.max_column
	number_data = []
	for i in range(2, max_row+1):
		ob = {}
		for col in range(1, max_col+1):
			ob[sheet_obj.cell(row = 1, column = col).value] = sheet_obj.cell(row = i, column = col).value
		number_data.append(ob);
	
	process_numbers(number_data)

def process_numbers(number_data):
	for user_record in number_data:
		if check_contact_availability(user_record) == True:
			select_contact()
			print(user_record)
			send_message(user_record)
			# check_sent_message_status()
		else:
			print(user_record)
			print("{}{}".format(user_record['number'], ' not available in your contacts'))
			send_message_to_unknown_contact(user_record)
	# logout()

def search_user_by_number(number):
	#input box to search within contacts
	print('searching ')
	print(number)
	name_search_input = driver.find_element_by_css_selector(".jN-F5.copyable-text.selectable-text")
	name_search_input.click()
	name_search_input.clear()
	name_search_input.send_keys(number)
	time.sleep(2)
	#after entering search number check if number found or not
	try:
		found_users_list = WebDriverWait(driver, 2).until(
        expected_conditions.presence_of_element_located((By.CSS_SELECTOR,"._2wP_Y ._2EXPL")))
		print('Contact found by numbber')
		return True
	except Exception as e:
		print(e)
		print("User not found in contacts");
		print("Will try to search with name...");
		return False;

def search_user_by_name(name):
	#input box to search within contacts
	num_search_input = driver.find_element_by_css_selector(".jN-F5.copyable-text.selectable-text")
	num_search_input.click()
	num_search_input.clear()
	num_search_input.send_keys(name)
	time.sleep(2)
	#after entering search number check if number found or not
	try:
		found_users_list = WebDriverWait(driver, 2).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR,"._2wP_Y ._2EXPL")))
		print('Contact found by name')
		return True
	except Exception as e:
		print(e)
		print("User not found in contacts");
		# print("Will try to search with name...");
		return False;
def send_message_to_unknown_contact(user_record):
	if user_record['type'] == 'text':
		url = "https://wa.me/91{}?text={}".format(user_record['number'],user_record['message'].replace("{#name#}",'User'))
		print(url)
		driver.get(url)
		# send button
		driver.find_element_by_css_selector("#action-button").click()
		# driver.find_element_by_css_selector("span[data-icon='send']").click();
	else:
		print('Cant not send images to unknown number')

def check_contact_availability(user_record):
	WebDriverWait(driver, 2).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR,".jN-F5.copyable-text.selectable-text")))
	if search_user_by_number(user_record['number']) == False:
		if user_record['name'] != '__BLANK__':
			if search_user_by_name(user_record['name']) == False:
				print(user_record['number'] + ' not available in your contacts.')
				return False
			else:
				return True
		else:
			return False
	else:
		return True

def select_contact():
	driver.find_element_by_css_selector("._2EXPL:not(._34xP_)").click()

def send_message(number_data):
	if number_data['type'] == 'text':
		#click on message input box
		driver.find_element_by_css_selector("#main > footer > div._3pkkz.copyable-area > div._1Plpp").click()
		#enter the message provided
		message_input = driver.find_element_by_css_selector("#main > footer > div._3pkkz.copyable-area > div._1Plpp")
		actions = ActionChains(driver)
		actions.move_to_element(message_input)
		actions.click(message_input)
		# get name of user
		users_display_name = driver.find_element_by_css_selector("#main > header > div._1WBXd > div > div > span").text
		actions.send_keys(number_data['message'].replace("{#name#}",users_display_name))
		actions.perform()
		#click send button
		driver.find_element_by_css_selector("._35EW6").click()
	elif number_data['type'] == 'image':
		# click attach button
		driver.find_element_by_css_selector("div[title='Attach']").click();
		# make element to be visible
		driver.execute_script("document.querySelectorAll('input[type=\"file\"]')[0].style.display = 'block'")
		driver.find_elements_by_css_selector("input[type='file']")[0].send_keys(os.path.realpath('./images/'+number_data['message']))
		
		# click send button 
		driver.find_element_by_css_selector("span[data-icon='send-light']").click();

@app.route("/")
def apiRunning():
	return "<h1>Working...</h1>"
	
@app.route('/whatsappmated/<path:path>')
def serve_files(path):
	print("******"+path)
	return send_from_directory('./', path)
	
if __name__ == "__main__":
    app.run(debug=True,host="0.0.0.0")


firebaseSetup()