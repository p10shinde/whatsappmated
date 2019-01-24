from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import os
from selenium.webdriver.firefox.options import Options
import base64
# from google.cloud import firestore
import uuid
from threading import Timer
import time
import openpyxl 


# db = firestore.Client()

mac_id = ':'.join(['{:02x}'.format((uuid.getnode() >> ele) & 0xff) 
for ele in range(0,8*6,8)][::-1])
'''
def hello():
	print('Checking qrCode')
	try:
		#check if user has logged in 
		#if this div._3dqpi presents then user has logged in else not scan qr code
		driver.get_screenshot_as_file("capture.png")
		logged_in_div = driver.find_element_by_css_selector("div._3dqpi")

		#search for the given number in list
		
		num_search_input = driver.find_element_by_css_selector(".jN-F5")
		num_search_input.send_keys("7355818205")

		#click first item found in the list
		driver.find_element_by_css_selector("._2wP_Y").find_element_by_css_selector("._2EXPL").click();

		#click message_input box
		driver.find_element_by_css_selector("#main > footer > div._3pkkz.copyable-area > div._1Plpp").click();
		print('element clicked')

		time.sleep(2);
		print('Typing message')
		#enter the message provided
		message_input = driver.find_element_by_css_selector("#main > footer > div._3pkkz.copyable-area > div._1Plpp")
		actions = ActionChains(driver)
		actions.move_to_element(message_input)
		actions.click(message_input)
		actions.send_keys("Hello userrrr")
		actions.perform()
		
		#click send button
		driver.find_element_by_css_selector("._35EW6").click();



		#if Image
		#click attach button
		driver.find_element_by_css_selector("div[title='Attach']").click();

		#make element to be visible
		driver.execute_script("document.querySelectorAll('input[type=\"file\"]')[0].style.display = 'block'")

		# print('selecting image')
		# driver.find_elements_by_css_selector("input[type='file']")[0].send_keys("C:\\xampp\\htdocs\\python\\Screenshot (1).png");
		driver.find_elements_by_css_selector("input[type='file']")[0].send_keys("http://www.tompetty.com/sites/g/files/g2000007521/f/sample001.jpg");
		
		# click send button 
		driver.find_element_by_css_selector("span[data-icon='send-light']").click();

		#image select button
		# driver.find_elements_by_css_selector("._10anr button")[0].click();





		# driver.get("https://web.whatsapp.com/send?phone=917355818205")
		print('User logged in')
		#continue to send message
	except:
		print('User not logged in')
	# try:
		qr_image_tag = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "._2EZ_m img")))
		qr_base64 = qr_image_tag.get_attribute("src")
	# except:
		# print("Timeout(20 seconds) : Could not load qr_code");
		# driver.quit()

		driver.get_screenshot_as_file("capture2.png")
		# qr_image_tag = driver.find_element_by_css_selector("._2EZ_m img")
		
		doc_ref = db.collection(u'data').document(mac_id)

		doc_ref.set({
		    'qr_code': qr_base64,
		    'name' : doc_ref.get().get('name')
		})

		print('qr_code updated')
		Timer(20.0, hello).start()
'''

# options = Options()
# # options.headless = True
# driver = webdriver.Firefox(options=options, executable_path="geckodriver.exe")
# driver.implicitly_wait(10)
# print("Firefox Headless Browser Invoked")
# driver.get('https://web.whatsapp.com')
# hello()
# Timer(20.0, hello).start()

# driver.quit()
# driver.get_screenshot_as_file("capture.png")

def setup_browser():
	options = Options()
	options.headless = True
	global driver
	driver = webdriver.Firefox(options=options, executable_path="geckodriver.exe")
	driver.implicitly_wait(10)
	print("Firefox Headless Browser Invoked")
	driver.get('https://web.whatsapp.com')
	return driver

def scan_qrcode(driver):
	try:
		#qr_code image
		qr_image_tag = WebDriverWait(driver, 15).until(
		expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "._2EZ_m img")))
		qr_base64 = qr_image_tag.get_attribute("src")
		driver.get_screenshot_as_file("qr_code.png")
		print('******qr scanned')
		return True
	except Exception as e: 
		print('scan qr error')
		print(e)
		return False
		# check_if_user_logged_in(driver)

def check_if_user_logged_in(driver):
	try:
		print('checking if user logged in?')
		#if this div presents then user is logged in
		logged_in_div = WebDriverWait(driver, 3).until(
        expected_conditions.presence_of_element_located((By.CSS_SELECTOR,"div._3dqpi")))
		return True
	except Exception as e: 
		# print('check_if_user_logged_in Error');
		# print(e)
		print("Yet user not logged in. Need to upadate qr_code")
		# print('waiting 12 sec')
		# time.sleep(12)
		# print('waiting 12 sec done')
		print('Scanning qr again')
		return False
		# scan_qrcode(driver)

def search_user_by_number(number):
	#input box to search within contacts
	print('searching ')
	print(number)
	name_search_input = driver.find_element_by_css_selector(".jN-F5.copyable-text.selectable-text")
	name_search_input.click()
	name_search_input.clear()
	name_search_input.send_keys(number)
	time.sleep(2)

	#after entering search number check if number found or not
	try:
		found_users_list = WebDriverWait(driver, 2).until(
        expected_conditions.presence_of_element_located((By.CSS_SELECTOR,"._2wP_Y ._2EXPL")))
		print('Contact found by numbber')
		return True
	except Exception as e:
		print(e)
		print("User not found in contacts");
		print("Will try to search with name...");
		return False;

def search_user_by_name(name):
	#input box to search within contacts
	num_search_input = driver.find_element_by_css_selector(".jN-F5.copyable-text.selectable-text")
	name_search_input.click()
	num_search_input.clear()
	num_search_input.send_keys(name)
	time.sleep(2)

	#after entering search number check if number found or not
	try:
		found_users_list = WebDriverWait(driver, 2).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR,"._2wP_Y ._2EXPL")))
		print('Contact found by name')
		return True
	except Exception as e:
		print(e)
		print("User not found in contacts");
		# print("Will try to search with name...");
		return False;

def check_contact_availability(user_record):
	
	WebDriverWait(driver, 2).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR,".jN-F5.copyable-text.selectable-text")))
	if search_user_by_number(user_record['number']) == False:
		if search_user_by_name(user_record['name']) == False:
			print(user_record['number'] + ' not available in your contacts.')
		else:
			return True
	else:
		return True

def select_contact():
	# driver.find_element_by_css_selector("._2wP_Y ._2EXPL").click()
	driver.find_element_by_css_selector("._2EXPL:not(._34xP_)").click()

def send_message(number_data):
	if number_data['type'] == 'text':
		#click on message input box
		driver.find_element_by_css_selector("#main > footer > div._3pkkz.copyable-area > div._1Plpp").click()
		#enter the message provided
		message_input = driver.find_element_by_css_selector("#main > footer > div._3pkkz.copyable-area > div._1Plpp")
		actions = ActionChains(driver)
		actions.move_to_element(message_input)
		actions.click(message_input)
		# get name of user
		users_display_name = driver.find_element_by_css_selector("#main > header > div._1WBXd > div > div > span").text
		actions.send_keys(number_data['message'].replace("{#name#}",users_display_name))
		actions.perform()
		#click send button
		driver.find_element_by_css_selector("._35EW6").click()
	elif number_data['type'] == 'image':
		# click attach button
		driver.find_element_by_css_selector("div[title='Attach']").click();
		# make element to be visible
		driver.execute_script("document.querySelectorAll('input[type=\"file\"]')[0].style.display = 'block'")
		driver.find_elements_by_css_selector("input[type='file']")[0].send_keys(os.path.realpath('./images/'+number_data['message']))
		
		# click send button 
		driver.find_element_by_css_selector("span[data-icon='send-light']").click();

def logout():
	#3 dot button for logout dropdown
	driver.find_element_by_css_selector("#side div:nth-child(3) > div[title='Menu']").click()

	

	#logout button
	driver.find_element_by_css_selector("div[title='Log out']").click()
	# after clicking logout wait for it to go to qr scanning page
	qr_image_tag = WebDriverWait(driver, 15).until(
	expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "._2EZ_m img")))
	print('User Logged out')
	driver.quit()
	

def check_sent_message_status():
	# msg-time --> when message sending is pending due to no network
	# msg-check --> when message is sent but not delivered
	# msg-dblcheck --> when message is sent and delivered but not red
	# msg-dblcheck-ack --> when message is sent, delivered and red

	all_visible_messages = driver.find_elements_by_css_selector("._32uRw span")
	if_any_msg_pending = False
	# visible_message = all_visible_messages[len(all_visible_messages)-1]
	# for visible_message in all_visible_messages:
	# print(visible_message.get_attribute('data-icon'))

	last_visible_message = all_visible_messages[len(all_visible_messages)-1]
	last_message_sent_status = "msg-time"
	while last_message_sent_status == "msg-time":
		time.sleep(.5)
		last_message_sent_status = last_visible_message.get_attribute('data-icon')

	print('Message sent')

def process_numbers(number_data):
	driver = setup_browser()
	if_qr_scan_success = False
	if_user_logged_in = False
	while if_qr_scan_success == False or if_user_logged_in == False:
		print('checking')
		if_qr_scan_success = scan_qrcode(driver)
		print('file_update')
		if_user_logged_in = check_if_user_logged_in(driver)
		if if_user_logged_in:
			if_qr_scan_success = True
		print(if_qr_scan_success)
		print(if_user_logged_in)







	# if scan_qrcode(driver) == False
	# 	print('waiting')
	# 	time.sleep(5)
	# 	print('waiting done')
	# 	if check_if_user_logged_in(driver) == False
	# 		scan_qrcode(driver)
	# 	else:
	print('user logged in')
	for user_record in number_data:
		if check_contact_availability(user_record) == True:
			select_contact()
			print(user_record)
			send_message(user_record)
			# check_sent_message_status()
		else:
			print(number_data['number'] + ' not available in your contacts')
	# logout()

def get_data_from_excel():
	wb_obj = openpyxl.load_workbook("./input/data.xlsx")
	sheet_obj = wb_obj.active
	max_row = sheet_obj.max_row
	max_col = sheet_obj.max_column
	number_data = []
	for i in range(2, max_row+1):
		ob = {}
		for col in range(1, max_col+1):
			ob[sheet_obj.cell(row = 1, column = col).value] = sheet_obj.cell(row = i, column = col).value
		number_data.append(ob);
	
	process_numbers(number_data)

get_data_from_excel()





